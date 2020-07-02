import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook((file))
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook((file))
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)


def readData(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook((file))
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetName, rownum, columnnum, data):
    workbook = openpyxl.load_workbook((file))
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)
