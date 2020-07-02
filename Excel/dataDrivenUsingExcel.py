import openpyxl

path = '/Users/govind794/Documents/GitHub/SeleniumAutomation/driver/Testing.xlsx'

workbook = openpyxl.load_workbook(path)

sheet = workbook.active
# sheet = workhook.get_sheet_by_name

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(row=r, column=c).value, end="       ")
    print()
