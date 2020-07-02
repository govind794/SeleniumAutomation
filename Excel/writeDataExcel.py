import openpyxl

path = '/Users/govind794/Documents/GitHub/SeleniumAutomation/driver/Testing1.xlsx'

workhook = openpyxl.load_workbook(path)

sheet = workhook.active

for r in range(1, 4):
    for c in range(1, 2):
        sheet.cell(row=r, column=c).value = 'welcome'
        # print(sheet.cell(row=r, column=c).value)
workhook.save(path)
 